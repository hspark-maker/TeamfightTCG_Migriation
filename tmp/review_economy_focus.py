from openpyxl import load_workbook
from pathlib import Path
import json
p=Path(r"C:\mgJeon\TeamfightTCG_Migriation\outputs\01a03695-cc18-7fe0-bdf5-bc71c5083e21\TeamfightTCG_경제시스템_설계.xlsx")
wf=load_workbook(p,data_only=False); wv=load_workbook(p,data_only=True)
for si in range(5,11):
 ws=wf.worksheets[si]; vs=wv[ws.title]
 print("SHEET",si,ws.title.encode('unicode_escape').decode(),ws.max_row,ws.max_column)
 for r in range(1,ws.max_row+1):
  row=[]
  for c in range(1,ws.max_column+1):
   x=ws.cell(r,c)
   if x.value is not None:
    v=str(x.value).encode('unicode_escape').decode()
    cv=vs.cell(r,c).value if x.data_type=='f' else None
    row.append(f"{x.coordinate}:{v}"+(f"=>{cv}" if x.data_type=='f' else ""))
  if row: print(" | ".join(row))
