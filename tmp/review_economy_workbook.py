from pathlib import Path
import json
from openpyxl import load_workbook

path = Path(r"C:\mgJeon\TeamfightTCG_Migriation\outputs\01a03695-cc18-7fe0-bdf5-bc71c5083e21\TeamfightTCG_경제시스템_설계.xlsx")
wb_f = load_workbook(path, data_only=False, read_only=False)
wb_v = load_workbook(path, data_only=True, read_only=False)

print("SHEETS", wb_f.sheetnames)
for ws in wb_f.worksheets:
    print(f"\n### {ws.title} {ws.max_row}x{ws.max_column} freeze={ws.freeze_panes} print={ws.print_area}")
    formulas = []
    errors = []
    for row in ws.iter_rows():
        for c in row:
            if c.data_type == "f":
                cached = wb_v[ws.title][c.coordinate].value
                formulas.append((c.coordinate, c.value, cached))
            if isinstance(c.value, str) and any(e in c.value for e in ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]):
                errors.append((c.coordinate, c.value))
    print("FORMULAS", len(formulas), "ERRORS", errors[:20])
    for item in formulas[:20]:
        print("F", *item)
    # Dump nonempty cells compactly for semantic review.
    for r in range(1, min(ws.max_row, 120) + 1):
        vals = []
        for c in range(1, min(ws.max_column, 20) + 1):
            cell = ws.cell(r, c)
            if cell.value is not None:
                val = str(cell.value).replace("\n", " / ")
                vals.append(f"{cell.coordinate}={val[:180]}")
        if vals:
            print("ROW", " | ".join(vals))

